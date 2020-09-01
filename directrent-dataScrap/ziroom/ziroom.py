#!/usr/bin/env python 
# -*- coding:utf-8 -*-
import requests
from lxml import etree
import time
import re
import os
import openpyxl


headers = {
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4103.106 Safari/537.36'
}

#获取所有房源的详情页url
def data_parse(url):
    res = requests.get(url=url,headers=headers)
    # print(res.text)
    demo = etree.HTML(res.text)
    try:
        data = demo.xpath('//div[@class="item"]/div[@class="pic-box"]/a/@href')
        for i in data:
            data_url.append(i)
    except:
        print('页面出错')
        print(res.text)

def detail_parse(url):
    res  = requests.get(url,headers=headers)
    html = etree.HTML(res.text)
    single_house_data = []
    global count
    count = count +1
    try:
        #图片url
        img_lists = html.xpath("//div[@id='Z_swiper_box']//ul[contains(@class,'Z_sliders_nav')]//li/img/@src")
        pics_path = download_pic(url,img_lists)
        #名字
        name = html.xpath("//aside[@class='Z_info_aside']/h1//text()")
        #标签
        tags = html.xpath("//div[@class='Z_tags']/span/text()")
        #面积
        area = html.xpath("//div[@class='Z_home_info']//dl[1]/dd/text()")
        #朝向
        towards = html.xpath("//div[@class='Z_home_info']//dl[2]/dd/text()")
        #户型
        unit_type = html.xpath("//div[@class='Z_home_info']//dl[3]/dd/text()")
        #位置
        location = html.xpath("//div[@class='Z_home_info']/ul[@class='Z_home_o']/li[1]//span[@class='ad']/text()")
        #楼层
        floor = html.xpath("//div[@class='Z_home_info']/ul[@class='Z_home_o']/li[2]//span[@class='va']/text()")
        #有无电梯
        if_elevator = html.xpath("//div[@class='Z_home_info']/ul[@class='Z_home_o']/li[3]//span[@class='va']/text()")
        #建成年份
        year = html.xpath("//div[@class='Z_home_info']/ul[@class='Z_home_o']/li[4]//span[@class='va']/text()")
        #门锁
        doorlock = html.xpath("//div[@class='Z_home_info']/ul[@class='Z_home_o']/li[5]//span[@class='va']/text()")
        #绿化
        greening = html.xpath("//div[@class='Z_home_info']/ul[@class='Z_home_o']/li[6]//span[@class='va']/text()")
        #房屋简介
        desc = html.xpath("//div[@class='Z_rent_desc']//text()")
        #配套设备(列表要除去更多和收起)
        equipment = html.xpath("//div[@class='Z_info_icons ']//dt/text()")
        for i in equipment:
            if i=='更多':
                equipment.remove(i)
            elif i=='收起':
                equipment.remove(i)
        #室友信息
        roommate = html.xpath("//div[@id='meetinfo']//li/div[@class='info']/p//text()")
        #去掉室友信息里得空字符和换行符
        new_roommate = []
        if roommate:
            for i in roommate:
                if i.strip():
                    new_roommate.append(i.strip())
        else:
            new_roommate.append("暂无信息")
        single_house_data.append(url)
        single_house_data.append(name[0])
        single_house_data.append(pics_path)
        single_house_data.append(tags)
        single_house_data.append(area[0])
        single_house_data.append(towards[0])
        single_house_data.append(unit_type[0])
        single_house_data.append(location[0])
        single_house_data.append(floor[0])
        single_house_data.append(if_elevator[0])
        single_house_data.append(year[0])
        single_house_data.append(doorlock[0])
        single_house_data.append(greening[0])
        if desc:
            if desc[0].strip():
                single_house_data.append(desc[0].strip())
            else:
                single_house_data.append('暂无描述')
        else:
            single_house_data.append('暂无描述')
        single_house_data.append(equipment)
        single_house_data.append(new_roommate)
        print('已获取第'+str(count)+'个房源数据')
        houses_data.append(single_house_data)

    except Exception as e:
        print(e)

def write_data(data):
    index = len(data)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "detail_data"
    sheet.cell(row=1, column=1, value='序号')
    sheet.cell(row=1, column=2, value='详情页')
    sheet.cell(row=1, column=3, value='名字')
    sheet.cell(row=1, column=4, value='图片文件夹名字')
    sheet.cell(row=1, column=5, value='房源标签')
    sheet.cell(row=1, column=6, value='面积')
    sheet.cell(row=1, column=7, value='朝向')
    sheet.cell(row=1, column=8, value='户型')
    sheet.cell(row=1, column=9, value='位置')
    sheet.cell(row=1, column=10, value='楼层')
    sheet.cell(row=1, column=11, value='有无电梯')
    sheet.cell(row=1, column=12, value='年份')
    sheet.cell(row=1, column=13, value='门锁')
    sheet.cell(row=1, column=14, value='绿化')
    sheet.cell(row=1, column=15, value='房源描述')
    sheet.cell(row=1, column=16, value='配置')
    sheet.cell(row=1, column=17, value='室友信息')
    for i in range(0, index):
        sheet.cell(row=i + 2, column=1, value=i + 1)
        sheet.cell(row=i + 2, column=2, value=data[i][0])
        sheet.cell(row=i + 2, column=3, value=data[i][1])
        sheet.cell(row=i + 2, column=4, value=data[i][2])
        sheet.cell(row=i + 2, column=5, value=','.join(data[i][3]))
        sheet.cell(row=i + 2, column=6, value=data[i][4])
        sheet.cell(row=i + 2, column=7, value=data[i][5])
        sheet.cell(row=i + 2, column=8, value=data[i][6])
        sheet.cell(row=i + 2, column=9, value=data[i][7])
        sheet.cell(row=i + 2, column=10, value=data[i][8])
        sheet.cell(row=i + 2, column=11, value=data[i][9])
        sheet.cell(row=i + 2, column=12, value=data[i][10])
        sheet.cell(row=i + 2, column=13, value=data[i][11])
        sheet.cell(row=i + 2, column=14, value=data[i][12])
        sheet.cell(row=i + 2, column=15, value=data[i][13])
        sheet.cell(row=i + 2, column=16, value=','.join(data[i][14]))
        sheet.cell(row=i + 2, column=17, value=','.join(data[i][15]))
    workbook.save("house_datas.xlsx")
    print("xlsx格式表格写入数据成功！")

def download_pic(detarl_url,pic_urls):
    #提取url里的数字作为文件夹名
    dir_name = re.findall(r"\d+?\d*", detarl_url)
    isExists = os.path.exists('./house_pics/'+dir_name[0])
    if not isExists:
        os.mkdir('./house_pics/'+dir_name[0])
    index = 0
    for i in pic_urls:
        index = index + 1
        if i.startswith("http"):
            res = requests.get(i, headers=headers)
            with open("./house_pics/" + dir_name[0] + "/" + str(index) + ".jpg", "wb") as f:
                f.write(res.content)
        else:
            res = requests.get('http:'+i, headers=headers)
            with open("./house_pics/" + dir_name[0] + "/" + str(index) + ".jpg", "wb") as f:
                f.write(res.content)

    return dir_name[0]

if __name__ == '__main__':
    data_url = []
    houses_data = []
    count = 0
    for i in range(1,51):
        url = 'http://cd.ziroom.com/z/p'+ str(i) + '/'
        data_parse(url)
    for i in data_url:
        if i.startswith('//cd.ziroom.com'):
            print(i)
            detail_parse('http:'+i)
    write_data(houses_data)

