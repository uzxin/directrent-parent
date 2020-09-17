#!/usr/bin/env python 
# -*- coding:utf-8 -*-
import xlrd
import price_spider
import openpyxl

def open_excel(file= 'house_datas.xlsx'):
    try:
        data = xlrd.open_workbook(file)
        return data
    except Exception as e:
        print(str(e))

def excel_table_byindex(file= 'house_datas.xlsx'):
    data = open_excel(file)
    table = data.sheets()[0]
    nrows = table.nrows #行数
    ncols = table.ncols #列数
    detail_urls =  table.col_values(1)
    detail_urls.pop(0)
    prices = price_spider.get_price(detail_urls)
    write_data(detail_urls,prices)

def write_data(detail_urls,prices):
    index = len(detail_urls)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "price"
    sheet.cell(row=1, column=1, value='序号')
    sheet.cell(row=1, column=2, value='详情页')
    sheet.cell(row=1, column=3, value='价格')
    for i in range(0, index):
        sheet.cell(row=i + 2, column=1, value=i + 1)
        sheet.cell(row=i + 2, column=2, value=detail_urls[i])
        sheet.cell(row=i + 2, column=3, value=prices[i])
    workbook.save("house_price.xlsx")
    print("xlsx格式表格写入数据成功！")


if __name__ == '__main__':
    excel_table_byindex()